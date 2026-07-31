#!/usr/bin/env python3

"""
Create MySQL tables from an Excel configuration workbook.

Input:
    JSON payload through stdin.

Excel format:
    Column A: project_id
    Column B: sub_project_id
    Remaining columns:
        First row of each pair  -> column headers
        Second row of each pair -> data types

This script creates tables only. It does not insert Excel records.
"""

from __future__ import annotations

import json
import os
import re
import sys
import traceback
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
import pymysql
from openpyxl.utils.datetime import from_excel
from pymysql.connections import Connection

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(
        encoding="utf-8",
        errors="replace",
        line_buffering=True,
    )

if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(
        encoding="utf-8",
        errors="replace",
        line_buffering=True,
    )


MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
MAX_TABLE_NAME_LENGTH = 64
MAX_COLUMN_NAME_LENGTH = 64
MAX_DYNAMIC_COLUMNS = 500

SUPPORTED_EXCEL_EXTENSION = ".xlsx"

PROJECT_ID_HEADER = "project_id"
SUB_PROJECT_ID_HEADER = "sub_project_id"
STATUS_MAPPING = {
    "AR Inprocess": "CE_Inprocess",
    "AR Pending": "CE_Pending",
    "AR Completed": "CE_Completed",
    "AR Hold": "CE_Hold",
    "QA Inprocess": "QA_Inprocess",
    "QA Pending": "QA_Pending",
    "QA Completed": "QA_Completed",
    "QA Hold": "QA_Hold",
    "AR Assigned": "CE_Assigned",
    "AR Non Workable": "AR_non_workable",
    "Auto Close": "Auto_Close",
}


SYSTEM_COLUMN_NAMES = {
    "id",
    "created_at",
    "updated_at",
    "deleted_at",
    "emp_id",
    "work_date",
    "charge_status",
    "invoke_date"
}

# ---------------------------------------------------------------------
# Supported logical datatype mappings
# ---------------------------------------------------------------------

TYPE_ALIASES = {
    "text": "TEXT",
    "string": "TEXT",
    "varchar": "TEXT",
    "date": "DATE",
    "datetime": "DATETIME",
    "timestamp": "DATETIME",
    "integer": "INT",
    "int": "INT",
    "whole number": "INT",
    "bigint": "BIGINT",
    "decimal": "DECIMAL",
    "number": "DECIMAL",
    "numeric": "DECIMAL",
    "float": "DECIMAL",
    "double": "DECIMAL",
    "boolean": "BOOLEAN",
    "bool": "BOOLEAN",
    "yes/no": "BOOLEAN",
    "long text": "LONGTEXT",
    "longtext": "LONGTEXT",
    "json": "JSON",
}


@dataclass(frozen=True)
class ColumnDefinition:
    source_header: str
    column_name: str
    mysql_type: str


@dataclass(frozen=True)
class TableDefinition:
    project_id: int
    sub_project_id: int
    project_name: str
    sub_project_name: str
    table_name: str
    columns: list[ColumnDefinition]
    header_excel_row: int
    datatype_excel_row: int


class ValidationError(Exception):
    """Expected workbook or business validation error."""


# ---------------------------------------------------------------------
# Response and logging helpers
# ---------------------------------------------------------------------

def log(message: str) -> None:
    """
    Diagnostic messages go to stderr so stdout remains valid JSON.
    """
    print(f"[dynamic-table] {message}", file=sys.stderr, flush=True)


def send_json(payload: dict[str, Any]) -> None:
    print(
        json.dumps(
            payload,
            ensure_ascii=True,
            default=str,
        ),
        flush=True,
    )


# ---------------------------------------------------------------------
# General normalization helpers
# ---------------------------------------------------------------------

def normalize_spaces(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_lookup_key(value: Any) -> str:
    return normalize_spaces(value).lower()


def normalize_identifier(value: Any, fallback_prefix: str) -> str:
    """
    Convert a project name, sub-project name or header into a safe
    lowercase MySQL identifier.
    """
    text = normalize_spaces(value)

    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()

    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text)
    text = text.strip("_")

    if not text:
        raise ValidationError(
            f"Unable to generate a valid identifier from '{value}'."
        )

    if text[0].isdigit():
        text = f"{fallback_prefix}_{text}"

    if len(text) > MAX_COLUMN_NAME_LENGTH:
        text = text[:MAX_COLUMN_NAME_LENGTH].rstrip("_")

    if not re.fullmatch(r"[a-z][a-z0-9_]*", text):
        raise ValidationError(
            f"Generated identifier '{text}' is invalid."
        )

    return text


def make_table_name(project_name: str, sub_project_name: str) -> str:
    project_slug = normalize_identifier(project_name, "project")
    sub_project_slug = normalize_identifier(
        sub_project_name,
        "subproject",
    )

    suffix = "_datas"
    base_name = f"{project_slug}_{sub_project_slug}"

    maximum_base_length = MAX_TABLE_NAME_LENGTH - len(suffix)
    base_name = base_name[:maximum_base_length].rstrip("_")

    table_name = f"{base_name}{suffix}"

    if not re.fullmatch(r"[a-z][a-z0-9_]*", table_name):
        raise ValidationError(
            f"Generated table name '{table_name}' is invalid."
        )

    return table_name


def parse_integer_id(value: Any, label: str, excel_row: int) -> int:
    if value is None or normalize_spaces(value) == "":
        raise ValidationError(
            f"{label} is missing at Excel row {excel_row}."
        )

    if isinstance(value, bool):
        raise ValidationError(
            f"{label} must be an integer at Excel row {excel_row}."
        )

    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValidationError(
            f"{label} must be an integer at Excel row {excel_row}."
        )

    if decimal_value != decimal_value.to_integral_value():
        raise ValidationError(
            f"{label} must be a whole number at Excel row {excel_row}."
        )

    parsed_value = int(decimal_value)

    if parsed_value <= 0:
        raise ValidationError(
            f"{label} must be greater than zero at Excel row {excel_row}."
        )

    return parsed_value


def map_mysql_type(
    datatype_value: Any,
    header: str,
    excel_row: int,
) -> str:
    datatype = normalize_lookup_key(datatype_value)

    if datatype == "":
        raise ValidationError(
            f"Datatype is missing for column '{header}' "
            f"at Excel row {excel_row}."
        )

    mapped_type = TYPE_ALIASES.get(datatype)

    if mapped_type is None:
        supported = ", ".join(sorted(TYPE_ALIASES.keys()))

        raise ValidationError(
            f"Unsupported datatype '{datatype_value}' for column "
            f"'{header}' at Excel row {excel_row}. "
            f"Supported values: {supported}."
        )

    if mapped_type == "TEXT":
        return "TEXT NULL"

    if mapped_type == "DATE":
        return "DATE NULL"

    if mapped_type == "DATETIME":
        return "DATETIME NULL"

    if mapped_type == "INT":
        return "INT NULL"

    if mapped_type == "BIGINT":
        return "BIGINT NULL"

    if mapped_type == "DECIMAL":
        return "DECIMAL(18,2) NULL"

    if mapped_type == "BOOLEAN":
        return "TINYINT(1) NULL"

    if mapped_type == "LONGTEXT":
        return "LONGTEXT NULL"

    if mapped_type == "JSON":
        return "JSON NULL"

    raise ValidationError(
        f"Internal datatype mapping is missing for '{datatype_value}'."
    )


# ---------------------------------------------------------------------
# Input and file security
# ---------------------------------------------------------------------

def read_input() -> dict[str, Any]:
    raw_input = sys.stdin.read()

    if not raw_input.strip():
        raise ValidationError("No input was received from Laravel.")

    try:
        payload = json.loads(raw_input)
    except json.JSONDecodeError as exception:
        raise ValidationError(
            "Laravel sent invalid JSON to the Python process."
        ) from exception

    if not isinstance(payload, dict):
        raise ValidationError("Input payload must be a JSON object.")

    return payload


def validate_uploaded_file(payload: dict[str, Any]) -> Path:
    file_path_value = payload.get("file_path")
    allowed_directory_value = payload.get("allowed_upload_directory")

    if not file_path_value:
        raise ValidationError("Uploaded file path is missing.")

    if not allowed_directory_value:
        raise ValidationError("Allowed upload directory is missing.")

    file_path = Path(str(file_path_value)).resolve()
    allowed_directory = Path(
        str(allowed_directory_value)
    ).resolve()

    if not file_path.is_file():
        raise ValidationError("The uploaded Excel file was not found.")

    try:
        file_path.relative_to(allowed_directory)
    except ValueError as exception:
        raise ValidationError(
            "The uploaded file is outside the approved upload directory."
        ) from exception

    if file_path.suffix.lower() != SUPPORTED_EXCEL_EXTENSION:
        raise ValidationError(
            "Please upload only an XLSX Excel file."
        )

    file_size = file_path.stat().st_size

    if file_size == 0:
        raise ValidationError("The uploaded Excel file is empty.")

    if file_size > MAX_FILE_SIZE_BYTES:
        raise ValidationError(
            "The uploaded Excel file exceeds the 10 MB limit."
        )

    return file_path


# ---------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------

def required_environment(name: str) -> str:
    value = os.getenv(name)

    if value is None or value == "":
        raise ValidationError(
            f"Required database environment variable '{name}' is missing."
        )

    return value


def get_connection() -> Connection:
    port_value = os.getenv("DYNAMIC_DB_PORT", "3306")

    try:
        port = int(port_value)
    except ValueError as exception:
        raise ValidationError(
            "Database port configuration is invalid."
        ) from exception

    return pymysql.connect(
        host=required_environment("DYNAMIC_DB_HOST"),
        port=port,
        user=required_environment("DYNAMIC_DB_USERNAME"),
        password=os.getenv("DYNAMIC_DB_PASSWORD", ""),
        database=required_environment("DYNAMIC_DB_DATABASE"),
        charset=os.getenv("DYNAMIC_DB_CHARSET", "utf8mb4"),
        connect_timeout=20,
        read_timeout=120,
        write_timeout=120,
        autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
    )


def fetch_project_and_subproject(
    connection: Connection,
    project_id: int,
    sub_project_id: int,
) -> tuple[str, str]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT project_name
            FROM projects
            WHERE project_id = %s
            LIMIT 1
            """,
            (project_id,),
        )

        project = cursor.fetchone()

        if not project:
            raise ValidationError(
                f"Project ID {project_id} does not exist."
            )

        cursor.execute(
            """
            SELECT sub_project_name
            FROM subprojects
            WHERE sub_project_id = %s
              AND project_id = %s
            LIMIT 1
            """,
            (sub_project_id, project_id),
        )

        sub_project = cursor.fetchone()

        if not sub_project:
            raise ValidationError(
                f"Sub-project ID {sub_project_id} does not belong "
                f"to project ID {project_id}."
            )

    project_name = normalize_spaces(project["project_name"])
    sub_project_name = normalize_spaces(
        sub_project["sub_project_name"]
    )

    if not project_name:
        raise ValidationError(
            f"Project name is empty for project ID {project_id}."
        )

    if not sub_project_name:
        raise ValidationError(
            f"Sub-project name is empty for sub-project ID "
            f"{sub_project_id}."
        )

    return project_name, sub_project_name


def validate_not_existing_in_ar_projects(
    connection: Connection,
    definitions: list[TableDefinition],
) -> None:
    """
    Prevent Non-AR table creation when the same active project and
    sub-project combination already exists in form_configurations.
    """

    with connection.cursor() as cursor:
        for definition in definitions:
            cursor.execute(
                """
                SELECT id
                FROM form_configurations
                WHERE project_id = %s
                  AND sub_project_id = %s
                  AND deleted_at IS NULL
                LIMIT 1
                """,
                (
                    str(definition.project_id),
                    str(definition.sub_project_id),
                ),
            )

            existing_ar_configuration = cursor.fetchone()

            if existing_ar_configuration:
                raise ValidationError(
                    f"{definition.project_name}_{definition.sub_project_name} "
                    "already exists in AR projects, so it cannot be "
                    "created in Non-AR projects."
                )


# ---------------------------------------------------------------------
# Workbook parsing
# ---------------------------------------------------------------------

def worksheet_has_content(worksheet: Any) -> bool:
    for row in worksheet.iter_rows(values_only=True):
        if any(
            value is not None and normalize_spaces(value) != ""
            for value in row
        ):
            return True

    return False


def validate_first_row(worksheet: Any) -> None:
    project_header = normalize_lookup_key(
        worksheet.cell(row=1, column=1).value
    )

    sub_project_header = normalize_lookup_key(
        worksheet.cell(row=1, column=2).value
    )

    if project_header != PROJECT_ID_HEADER:
        raise ValidationError(
            "Excel cell A1 must contain 'project_id'."
        )

    if sub_project_header != SUB_PROJECT_ID_HEADER:
        raise ValidationError(
            "Excel cell B1 must contain 'sub_project_id'."
        )


def row_is_empty(
    worksheet: Any,
    row_number: int,
) -> bool:
    for column_number in range(1, worksheet.max_column + 1):
        value = worksheet.cell(
            row=row_number,
            column=column_number,
        ).value

        if value is not None and normalize_spaces(value) != "":
            return False

    return True


def build_column_definitions(
    worksheet: Any,
    header_row: int,
    datatype_row: int,
) -> list[ColumnDefinition]:
    columns: list[ColumnDefinition] = []
    used_column_names: dict[str, str] = {}

    for column_number in range(3, worksheet.max_column + 1):
        raw_header = worksheet.cell(
            row=header_row,
            column=column_number,
        ).value

        raw_datatype = worksheet.cell(
            row=datatype_row,
            column=column_number,
        ).value

        header = normalize_spaces(raw_header)
        datatype = normalize_spaces(raw_datatype)

        if header == "" and datatype == "":
            continue

        excel_column = openpyxl.utils.get_column_letter(
            column_number
        )

        if header == "":
            raise ValidationError(
                f"Header is missing in Excel column {excel_column}, "
                f"row {header_row}, but a datatype was provided."
            )

        if datatype == "":
            raise ValidationError(
                f"Datatype is missing for header '{header}' in "
                f"Excel column {excel_column}, row {datatype_row}."
            )

        column_name = normalize_identifier(header, "column")

        if column_name in SYSTEM_COLUMN_NAMES:
            raise ValidationError(
                f"Header '{header}' generates reserved column name "
                f"'{column_name}'. This column is created automatically."
            )

        if column_name in used_column_names:
            previous_header = used_column_names[column_name]

            raise ValidationError(
                f"Headers '{previous_header}' and '{header}' both "
                f"generate the same database column '{column_name}'."
            )

        used_column_names[column_name] = header

        mysql_type = map_mysql_type(
            raw_datatype,
            header,
            datatype_row,
        )

        columns.append(
            ColumnDefinition(
                source_header=header,
                column_name=column_name,
                mysql_type=mysql_type,
            )
        )

    if not columns:
        raise ValidationError(
            f"No table columns were found in Excel row {header_row}."
        )

    if len(columns) > MAX_DYNAMIC_COLUMNS:
        raise ValidationError(
            f"Excel row {header_row} contains more than "
            f"{MAX_DYNAMIC_COLUMNS} dynamic columns."
        )

    return columns


def parse_workbook(
    file_path: Path,
    connection: Connection,
) -> list[TableDefinition]:
    try:
        workbook = openpyxl.load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )
    except Exception as exception:
        raise ValidationError(
            "The XLSX file is invalid, corrupted or password protected."
        ) from exception

    try:
        non_empty_worksheets = [
            worksheet
            for worksheet in workbook.worksheets
            if worksheet_has_content(worksheet)
        ]

        if not non_empty_worksheets:
            raise ValidationError(
                "The Excel workbook does not contain any data."
            )

        if len(non_empty_worksheets) > 1:
            sheet_names = ", ".join(
                worksheet.title
                for worksheet in non_empty_worksheets
            )

            raise ValidationError(
                "Only one populated worksheet is allowed. "
                f"Populated worksheets found: {sheet_names}."
            )

        worksheet = non_empty_worksheets[0]

        validate_first_row(worksheet)

        definitions: list[TableDefinition] = []
        encountered_combinations: set[tuple[int, int]] = set()
        encountered_table_names: set[str] = set()

        row_number = 2

        while row_number <= worksheet.max_row:
            if row_is_empty(worksheet, row_number):
                row_number += 1
                continue

            header_row = row_number
            datatype_row = row_number + 1

            if datatype_row > worksheet.max_row:
                raise ValidationError(
                    f"Excel row {header_row} contains headers but its "
                    "datatype row is missing."
                )

            if row_is_empty(worksheet, datatype_row):
                raise ValidationError(
                    f"Datatype row {datatype_row} is empty for "
                    f"header row {header_row}."
                )

            header_project_id = parse_integer_id(
                worksheet.cell(
                    row=header_row,
                    column=1,
                ).value,
                "project_id",
                header_row,
            )

            header_sub_project_id = parse_integer_id(
                worksheet.cell(
                    row=header_row,
                    column=2,
                ).value,
                "sub_project_id",
                header_row,
            )

            datatype_project_id = parse_integer_id(
                worksheet.cell(
                    row=datatype_row,
                    column=1,
                ).value,
                "project_id",
                datatype_row,
            )

            datatype_sub_project_id = parse_integer_id(
                worksheet.cell(
                    row=datatype_row,
                    column=2,
                ).value,
                "sub_project_id",
                datatype_row,
            )

            if header_project_id != datatype_project_id:
                raise ValidationError(
                    f"Project ID mismatch between Excel rows "
                    f"{header_row} and {datatype_row}."
                )

            if header_sub_project_id != datatype_sub_project_id:
                raise ValidationError(
                    f"Sub-project ID mismatch between Excel rows "
                    f"{header_row} and {datatype_row}."
                )

            combination = (
                header_project_id,
                header_sub_project_id,
            )

            if combination in encountered_combinations:
                raise ValidationError(
                    f"Project ID {header_project_id} and sub-project "
                    f"ID {header_sub_project_id} appear more than once "
                    "in the workbook."
                )

            project_name, sub_project_name = (
                fetch_project_and_subproject(
                    connection,
                    header_project_id,
                    header_sub_project_id,
                )
            )

            table_name = make_table_name(
                project_name,
                sub_project_name,
            )

            if table_name in encountered_table_names:
                raise ValidationError(
                    f"More than one Excel configuration generates the "
                    f"same table name '{table_name}'."
                )

            columns = build_column_definitions(
                worksheet,
                header_row,
                datatype_row,
            )

            definitions.append(
                TableDefinition(
                    project_id=header_project_id,
                    sub_project_id=header_sub_project_id,
                    project_name=project_name,
                    sub_project_name=sub_project_name,
                    table_name=table_name,
                    columns=columns,
                    header_excel_row=header_row,
                    datatype_excel_row=datatype_row,
                )
            )

            encountered_combinations.add(combination)
            encountered_table_names.add(table_name)

            row_number += 2

        if not definitions:
            raise ValidationError(
                "No valid table configurations were found in Excel."
            )

        return definitions

    finally:
        workbook.close()


# ---------------------------------------------------------------------
# Existing schema validation
# ---------------------------------------------------------------------
def normalize_existing_mysql_type(
    column: dict[str, Any],
) -> str:
    """
    Normalize MySQL/MariaDB column types for schema comparison.

    Examples:
        BIGINT(20) UNSIGNED -> BIGINT UNSIGNED
        INT(11)             -> INT
        TINYINT(1)          -> TINYINT(1)
        VARCHAR(255)        -> VARCHAR(255)
        ENUM(...)           -> ENUM(...)
    """

    data_type = normalize_lookup_key(
        column.get("DATA_TYPE")
    )

    column_type = normalize_spaces(
        column.get("COLUMN_TYPE")
    ).upper()

    if data_type in {
        "bigint",
        "int",
        "integer",
        "smallint",
        "mediumint",
    }:
        # Remove MariaDB/MySQL integer display width.
        # BIGINT(20) UNSIGNED becomes BIGINT UNSIGNED.
        column_type = re.sub(
            r"\(\d+\)",
            "",
            column_type,
        )

        column_type = re.sub(
            r"\s+",
            " ",
            column_type,
        ).strip()

        return column_type

    if data_type == "tinyint":
        # Preserve TINYINT(1), since it is commonly used for boolean.
        return column_type

    if data_type in {
        "varchar",
        "char",
        "decimal",
        "numeric",
        "enum",
        "set",
    }:
        return column_type

    return data_type.upper()
def normalize_expected_mysql_type(
    mysql_type: str,
) -> str:
    normalized_type = re.sub(
        r"\s+NULL$",
        "",
        mysql_type.strip(),
        flags=re.IGNORECASE,
    )

    return normalized_type.upper()


def table_exists(
    connection: Connection,
    table_name: str,
) -> bool:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT COUNT(*) AS table_count
            FROM information_schema.TABLES
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = %s
            """,
            (table_name,),
        )

        result = cursor.fetchone()

    return int(result["table_count"]) > 0


def get_existing_columns(
    connection: Connection,
    table_name: str,
) -> list[dict[str, Any]]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                COLUMN_NAME,
                DATA_TYPE,
                COLUMN_TYPE,
                IS_NULLABLE,
                COLUMN_KEY,
                EXTRA,
                ORDINAL_POSITION
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = %s
            ORDER BY ORDINAL_POSITION
            """,
            (table_name,),
        )

        return list(cursor.fetchall())
def parse_mysql_enum_values(column_type: str) -> set[str]:
    """
    Parse MySQL ENUM values returned by information_schema.

    Example:
        enum('CE_Assigned','CE_Inprocess')
    """

    if not column_type:
        return set()

    return set(
        re.findall(
            r"'((?:[^'\\\\]|\\\\.)*)'",
            column_type,
        )
    )
def validate_existing_system_column_types(
    table_name: str,
    existing_column_map: dict[str, dict[str, Any]],
    errors: list[str],
) -> None:
    """
    Validate automatically created system columns.
    """

    expected_types = {
        "id": "BIGINT UNSIGNED",
        "emp_id": "TEXT",
        "work_date": "DATE",
        "invoke_date": "DATE",
        "created_at": "TIMESTAMP",
        "updated_at": "TIMESTAMP",
        "deleted_at": "TIMESTAMP",
    }

    for column_name, expected_type in expected_types.items():
        column = existing_column_map.get(column_name)

        if not column:
            continue

        actual_type = normalize_existing_mysql_type(
            column
        )

        if actual_type != expected_type:
            errors.append(
                f"system column '{column_name}' expects "
                f"{expected_type} but existing type is {actual_type}"
            )

    charge_status_column = existing_column_map.get(
        "charge_status"
    )

    if not charge_status_column:
        return

    actual_charge_status_type = normalize_existing_mysql_type(
        charge_status_column
    )

    if not actual_charge_status_type.startswith("ENUM("):
        errors.append(
            "system column 'charge_status' must be an ENUM"
        )
        return

    expected_status_values = set(
        STATUS_MAPPING.values()
    )

    actual_status_values = parse_mysql_enum_values(
        str(
            charge_status_column.get(
                "COLUMN_TYPE",
                ""
            )
        )
    )

    missing_status_values = (
        expected_status_values - actual_status_values
    )

    if missing_status_values:
        errors.append(
            "charge_status is missing enum values: "
            + ", ".join(
                sorted(missing_status_values)
            )
        )
def validate_existing_table(
    connection: Connection,
    definition: TableDefinition,
) -> None:
    """
    Validate an existing dynamic table.

    Excel-defined columns and automatically created system columns
    are validated separately.
    """

    existing_columns = get_existing_columns(
        connection,
        definition.table_name,
    )

    expected_dynamic_columns = {
        column.column_name.lower():
            normalize_expected_mysql_type(
                column.mysql_type
            )
        for column in definition.columns
    }

    # Normalize column names for case-insensitive comparison.
    existing_column_map = {
        str(column["COLUMN_NAME"]).lower(): column
        for column in existing_columns
    }

    required_system_columns = {
        "id",
        "emp_id",
        "work_date",
        "charge_status",
        "invoke_date",
        "created_at",
        "updated_at",
        "deleted_at",
    }

    existing_column_names = set(
        existing_column_map.keys()
    )

    missing_system_columns = (
        required_system_columns - existing_column_names
    )

    if missing_system_columns:
        raise ValidationError(
            f"Table '{definition.table_name}' already exists but is "
            "not compatible. Missing system columns: "
            + ", ".join(sorted(missing_system_columns))
            + "."
        )

    existing_dynamic_names = (
        existing_column_names - required_system_columns
    )

    expected_dynamic_names = set(
        expected_dynamic_columns.keys()
    )

    missing_columns = (
        expected_dynamic_names - existing_dynamic_names
    )

    extra_columns = (
        existing_dynamic_names - expected_dynamic_names
    )

    errors: list[str] = []

    if missing_columns:
        errors.append(
            "missing Excel columns: "
            + ", ".join(sorted(missing_columns))
        )

    if extra_columns:
        errors.append(
            "extra columns: "
            + ", ".join(sorted(extra_columns))
        )

    for column_name in sorted(
        expected_dynamic_names & existing_dynamic_names
    ):
        actual_type = normalize_existing_mysql_type(
            existing_column_map[column_name]
        )

        expected_type = expected_dynamic_columns[
            column_name
        ]

        if actual_type != expected_type:
            errors.append(
                f"column '{column_name}' expects {expected_type} "
                f"but existing type is {actual_type}"
            )

    validate_existing_system_column_types(
        definition.table_name,
        existing_column_map,
        errors,
    )

    if errors:
        raise ValidationError(
            f"Table '{definition.table_name}' already exists with a "
            "different structure: "
            + "; ".join(errors)
            + ". Existing tables are not altered automatically."
        )
# ---------------------------------------------------------------------
# Table creation
# ---------------------------------------------------------------------

def quote_identifier(identifier: str) -> str:
    """
    Identifiers are already normalized, but this remains a second
    defensive layer.
    """
    if not re.fullmatch(r"[a-z][a-z0-9_]*", identifier):
        raise ValidationError(
            f"Unsafe database identifier '{identifier}'."
        )

    return f"`{identifier}`"

def build_create_table_sql(
    definition: TableDefinition,
) -> str:
    """
    Build CREATE TABLE SQL for the dynamic _datas table.
    """

    charge_status_values = list(
        dict.fromkeys(STATUS_MAPPING.values())
    )

    escaped_status_values = [
        "'" + value.replace("'", "''") + "'"
        for value in charge_status_values
    ]

    charge_status_enum = ",".join(
        escaped_status_values
    )

    column_sql = [
        "`id` BIGINT UNSIGNED NOT NULL AUTO_INCREMENT",
    ]

    for column in definition.columns:
        column_sql.append(
            f"{quote_identifier(column.column_name)} "
            f"{column.mysql_type}"
        )

    column_sql.extend([
        "`emp_id` TEXT NULL",
        "`work_date` DATE NULL",
        (
            "`charge_status` ENUM("
            + charge_status_enum
            + ") NULL DEFAULT 'CE_Completed'"
        ),
        "`invoke_date` DATE NULL",
        "`created_at` TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP",
        "`updated_at` TIMESTAMP NULL DEFAULT NULL",
        "`deleted_at` TIMESTAMP NULL DEFAULT NULL",
        "PRIMARY KEY (`id`)",
    ])

    return (
        f"CREATE TABLE {quote_identifier(definition.table_name)} "
        "(\n    "
        + ",\n    ".join(column_sql)
        + "\n) ENGINE=InnoDB "
        "DEFAULT CHARACTER SET utf8mb4 "
        "COLLATE utf8mb4_unicode_ci"
    )
def format_excel_header_for_configuration(header: str) -> str:
    """
    Format the original Excel header for data_columns/date_columns.

    Examples:
        "Unique ID"   -> "Unique_ID"
        "Patient No"  -> "Patient_No"
        "provider_no" -> "provider_no"
        "FC"          -> "FC"
        "AR Emp Id"   -> "AR_Emp_Id"

    Original capitalization is preserved.
    """

    value = str(header or "").strip()

    # Replace any spaces or unsupported characters with underscore.
    value = re.sub(
        r"[^A-Za-z0-9_]+",
        "_",
        value
    )

    # Replace multiple underscores with one underscore.
    value = re.sub(
        r"_+",
        "_",
        value
    )

    return value.strip("_")
def normalize_json_array_for_comparison(value: Any) -> list[str]:
    """
    Convert existing database configuration to a normalized Python list.

    Supports:
        JSON array string
        Comma-separated string
        Python list
        NULL or empty value
    """

    if value is None:
        return []

    if isinstance(value, list):
        return [
            str(item).strip()
            for item in value
            if str(item).strip()
        ]

    raw_value = str(value).strip()

    if raw_value == "":
        return []

    try:
        decoded_value = json.loads(raw_value)

        if isinstance(decoded_value, list):
            return [
                str(item).strip()
                for item in decoded_value
                if str(item).strip()
            ]

    except (json.JSONDecodeError, TypeError):
        pass

    return [
        item.strip()
        for item in raw_value.split(",")
        if item.strip()
    ]
def save_non_ar_upload_configuration(
    connection: Connection,
    definition: TableDefinition,
) -> str:

    system_columns = SYSTEM_COLUMN_NAMES

    data_columns = [
        format_excel_header_for_configuration(
            column.source_header
        )
        for column in definition.columns
    ]

    db_columns = [
        column.column_name
        for column in definition.columns
        if column.column_name not in system_columns
    ]

    date_columns = [
        format_excel_header_for_configuration(
            column.source_header
        )
        for column in definition.columns
        if normalize_expected_mysql_type(
            column.mysql_type
        ) in {
            "DATE",
            "DATETIME",
        }
    ]

    # Add Emp_Id mapping.
    # File/template column: Emp_Id
    # Dynamic table column: emp_id
    if not any(
        column.lower() == "emp_id"
        for column in data_columns
    ):
        data_columns.append("Emp_Id")

    if not any(
        column.lower() == "emp_id"
        for column in db_columns
    ):
        db_columns.append("emp_id")

    # Add Work_date mapping.
    # File/template column: Work_date
    # Dynamic table column: work_date
    if not any(
        column.lower() == "work_date"
        for column in data_columns
    ):
        data_columns.append("Work_date")

    if not any(
        column.lower() == "work_date"
        for column in db_columns
    ):
        db_columns.append("work_date")

    # Work_date must also be handled as a date column.
    if not any(
        column.lower() == "work_date"
        for column in date_columns
    ):
        date_columns.append("Work_date")

    data_columns_value = ",".join(data_columns)
    db_columns_value = ",".join(db_columns)
    date_columns_value = ",".join(date_columns)

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                id,
                table_name,
                data_columns,
                db_columns,
                date_columns,
                deleted_at
            FROM non_ar_inventory_upload_configuration
            WHERE project_id = %s
              AND sub_project_id = %s
            ORDER BY id ASC
            LIMIT 1
            """,
            (
                str(definition.project_id),
                str(definition.sub_project_id),
            ),
        )

        existing_configuration = cursor.fetchone()

        if not existing_configuration:
            cursor.execute(
                """
                INSERT INTO non_ar_inventory_upload_configuration
                (
                    project_id,
                    sub_project_id,
                    table_name,
                    data_columns,
                    db_columns,
                    required_columns,
                    date_columns,
                    numeric_columns,
                    duplicate_columns,
                    created_at,
                    updated_at,
                    deleted_at
                )
                VALUES
                (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    NULL,
                    %s,
                    NULL,
                    NULL,
                    NOW(),
                    NOW(),
                    NULL
                )
                """,
                (
                    str(definition.project_id),
                    str(definition.sub_project_id),
                    definition.table_name,
                    data_columns_value,
                    db_columns_value,
                    date_columns_value,
                ),
            )

            return "inserted"

        existing_data_columns = normalize_json_array_for_comparison(
            existing_configuration.get("data_columns")
        )

        existing_db_columns = normalize_json_array_for_comparison(
            existing_configuration.get("db_columns")
        )

        existing_date_columns = normalize_json_array_for_comparison(
            existing_configuration.get("date_columns")
        )

        configuration_changed = (
            str(existing_configuration.get("table_name") or "")
            != definition.table_name
            or existing_data_columns != data_columns
            or existing_db_columns != db_columns
            or existing_date_columns != date_columns
            or existing_configuration.get("deleted_at") is not None
        )

        if not configuration_changed:
            return "unchanged"

        cursor.execute(
            """
            UPDATE non_ar_inventory_upload_configuration
            SET table_name = %s,
                data_columns = %s,
                db_columns = %s,
                date_columns = %s,
                updated_at = NOW(),
                deleted_at = NULL
            WHERE id = %s
            """,
            (
                definition.table_name,
                data_columns_value,
                db_columns_value,
                date_columns_value,
                existing_configuration["id"],
            ),
        )

        return "updated"
def process_table_definitions(
    connection: Connection,
    definitions: list[TableDefinition],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []

    # Prevent Non-AR creation for active AR project configurations.
    validate_not_existing_in_ar_projects(
        connection,
        definitions,
    )

    # Validate all existing tables before processing.
    for definition in definitions:
        if table_exists(connection, definition.table_name):
            validate_existing_table(
                connection,
                definition,
            )

    try:
        for definition in definitions:
            exists = table_exists(
                connection,
                definition.table_name,
            )

            if exists:
                table_status = "already_exists"
            else:
                create_sql = build_create_table_sql(
                    definition
                )

                with connection.cursor() as cursor:
                    cursor.execute(create_sql)

                table_status = "created"

            # Insert or update the non-AR upload configuration.
            configuration_status = (
                save_non_ar_upload_configuration(
                    connection,
                    definition,
                )
            )

            results.append({
                "project_id": definition.project_id,
                "sub_project_id": definition.sub_project_id,
                "project_name": definition.project_name,
                "sub_project_name": definition.sub_project_name,
                "table": definition.table_name,
                "status": table_status,
                "configuration_status": configuration_status,
                "configuration_table": (
                    "non_ar_inventory_upload_configuration"
                ),
                "column_count": len(definition.columns),
                "columns": [
                    {
                        "excel_header": column.source_header,
                        "database_column": column.column_name,
                        "database_type": (
                            normalize_expected_mysql_type(
                                column.mysql_type
                            )
                        ),
                        "is_date_column": (
                            normalize_expected_mysql_type(
                                column.mysql_type
                            )
                            in {
                                "DATE",
                                "DATETIME",
                            }
                        ),
                    }
                    for column in definition.columns
                ],
                "header_excel_row": (
                    definition.header_excel_row
                ),
                "datatype_excel_row": (
                    definition.datatype_excel_row
                ),
            })

        connection.commit()

        return results

    except Exception:
        connection.rollback()
        raise


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main() -> int:
    connection: Connection | None = None

    try:
        payload = read_input()
        file_path = validate_uploaded_file(payload)

        connection = get_connection()

        definitions = parse_workbook(
            file_path,
            connection,
        )

        # Validate all project/sub-project combinations before creating
        # dynamic tables or saving Non-AR upload configuration records.
        validate_not_existing_in_ar_projects(
            connection,
            definitions,
        )

        results = process_table_definitions(
            connection,
            definitions,
        )

        created_count = sum(
            result["status"] == "created"
            for result in results
        )

        existing_count = sum(
            result["status"] == "already_exists"
            for result in results
        )

        send_json({
            "status": "success",
            "message": (
                f"{created_count} table(s) created and "
                f"{existing_count} compatible existing table(s) "
                "verified successfully."
            ),
            "data": {
                "processed": len(results),
                "created": created_count,
                "already_exists": existing_count,
                "tables": results,
            },
        })

        return 0

    except ValidationError as exception:
        log(str(exception))

        send_json({
            "status": "warning",
            "message": str(exception),
            "errors": [],
        })

        return 1

    except pymysql.MySQLError as exception:
        log(f"MySQL error: {exception}")

        send_json({
            "status": "warning",
            "message": (
                "Database operation failed while creating the "
                "dynamic tables."
            ),
            "errors": [
                {
                    "type": type(exception).__name__,
                    "detail": str(exception),
                }
            ],
        })

        return 1

    except Exception as exception:
        log(traceback.format_exc())

        send_json({
            "status": "warning",
            "message": (
                "Unexpected error occurred while processing the "
                "Excel configuration."
            ),
            "errors": [
                {
                    "type": type(exception).__name__,
                    "detail": str(exception),
                }
            ],
        })

        return 1

    finally:
        if connection is not None:
            connection.close()


if __name__ == "__main__":
    sys.exit(main())